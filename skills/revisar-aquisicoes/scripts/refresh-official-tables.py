from __future__ import annotations

import argparse
import hashlib
import json
import re
import tempfile
import urllib.request
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SOURCE_URL = (
    "https://www.nfe.fazenda.gov.br/portal/exibirArquivo.aspx?conteudo=D5b4Ov84WDg="
)
PORTAL_URL = (
    "https://www.nfe.fazenda.gov.br/portal/listaConteudo.aspx?tipoConteudo=/NJarYc9nus="
)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _iso(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value in {None, ""}:
        return None
    return str(value).strip() or None


def _flag(value: Any) -> int:
    return 1 if value in {1, "1"} else 0


def _decimal_text(value: Any) -> str | None:
    if value in {None, ""}:
        return None
    return str(value).strip()


def build_snapshot(source: Path) -> dict[str, Any]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    cst_sheet = workbook[workbook.sheetnames[0]]
    class_sheet = workbook[workbook.sheetnames[1]]

    cst_records: list[dict[str, Any]] = []
    for row in cst_sheet.iter_rows(min_row=2, values_only=True):
        cst = str(row[0] or "").strip()
        if re.fullmatch(r"\d{3}", cst) is None:
            continue
        cst_records.append(
            {
                "cst": cst,
                "ind_gIBSCBS": _flag(row[2]),
                "ind_gIBSCBSMono": _flag(row[3]),
                "ind_gRed": _flag(row[4]),
                "ind_gDif": _flag(row[5]),
                "ind_gTransfCred": _flag(row[6]),
                "ind_gCredPresIBSZFM": _flag(row[7]),
                "ind_gAjusteCompet": _flag(row[8]),
                "ind_RedutorBC": _flag(row[9]),
            }
        )

    classification_records: list[dict[str, Any]] = []
    for row in class_sheet.iter_rows(min_row=2, values_only=True):
        cst = str(row[0] or "").strip()
        cclass = str(row[2] or "").strip()
        if (
            re.fullmatch(r"\d{3}", cst) is None
            or re.fullmatch(r"\d{6}", cclass) is None
        ):
            continue
        classification_records.append(
            {
                "cst": cst,
                "cclass_trib": cclass,
                "name": str(row[3] or "").strip() or None,
                "lc214_reference": str(row[6] or "").strip() or None,
                "rate_type": str(row[9] or "").strip() or None,
                "ibs_reduction_percent": _decimal_text(row[10]),
                "cbs_reduction_percent": _decimal_text(row[11]),
                "effective_from": _iso(row[21]),
                "effective_to": _iso(row[22]),
                "updated_at": _iso(row[23]),
                "document_applicability": {
                    "NFE": _flag(row[25]),
                    "NFCE": _flag(row[26]),
                    "CTE": _flag(row[27]),
                    "NFSE": _flag(row[33]),
                },
                "legal_link": str(row[42] or "").strip() or None,
            }
        )

    return {
        "schema": "br.com.planejamento-reforma-tributaria/official-tax-snapshot",
        "schema_version": "1.0.0",
        "snapshot_id": "CCLASSTRIB-IT-2025.002-V1.60-2026-06-22",
        "verified_at": "2026-08-31",
        "source": {
            "title": "Tabela de Classificação Tributária do IBS e CBS",
            "technical_version": "IT 2025.002 v1.60",
            "publication_date": "2026-06-23",
            "file_name": "cClassTrib 2026-06-22.xlsx",
            "file_url": SOURCE_URL,
            "portal_url": PORTAL_URL,
            "xlsx_sha256": _sha256(source),
        },
        "legal_sources": [
            {
                "id": "LC-214-2025",
                "url": "https://www.planalto.gov.br/ccivil_03/leis/lcp/lcp214.htm",
            },
            {
                "id": "LC-227-2026",
                "url": "https://www.planalto.gov.br/ccivil_03/leis/lcp/lcp227.htm",
            },
            {
                "id": "DECRETO-12955-2026",
                "url": "https://www.planalto.gov.br/ccivil_03/_ato2023-2026/2026/decreto/d12955.htm",
            },
        ],
        "cst_records": sorted(cst_records, key=lambda item: item["cst"]),
        "classification_records": sorted(
            classification_records,
            key=lambda item: (item["cst"], item["cclass_trib"]),
        ),
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-xlsx", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    temporary: tempfile.TemporaryDirectory[str] | None = None
    source = args.input_xlsx
    if source is None:
        temporary = tempfile.TemporaryDirectory(prefix="rtc-official-")
        source = Path(temporary.name) / "cclass.xlsx"
        urllib.request.urlretrieve(SOURCE_URL, source)
    try:
        snapshot = build_snapshot(source.resolve())
        target = args.output.resolve()
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(
            json.dumps(snapshot, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )
        print(
            json.dumps(
                {
                    "snapshot_id": snapshot["snapshot_id"],
                    "cst_records": len(snapshot["cst_records"]),
                    "classification_records": len(snapshot["classification_records"]),
                    "output": str(target),
                },
                ensure_ascii=False,
                sort_keys=True,
            )
        )
    finally:
        if temporary is not None:
            temporary.cleanup()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
