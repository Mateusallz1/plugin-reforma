from __future__ import annotations

import argparse
import hashlib
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SOURCE_URL = (
    "https://www.nfe.fazenda.gov.br/portal/exibirArquivo.aspx?conteudo=4kyLWvNwhBY="
)
PORTAL_URL = (
    "https://www.nfe.fazenda.gov.br/portal/listaConteudo.aspx?tipoConteudo=/NJarYc9nus="
)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _iso(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value in {None, ""}:
        return None
    return str(value).strip() or None


def _flag(value: Any) -> int:
    return 1 if value in {1, "1"} else 0


def build_snapshot(source: Path) -> dict[str, Any]:
    sheet = load_workbook(source, read_only=True, data_only=True)["CFOP"]
    records: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, max_col=12, values_only=True):
        cfop = str(row[0] or "").strip()
        if re.fullmatch(r"\d{4}", cfop) is None:
            continue
        records.append(
            {
                "cfop": cfop,
                "effective_from": _iso(row[1]),
                "effective_to": _iso(row[2]),
                "ind_nfe": _flag(row[3]),
                "ind_communication": _flag(row[4]),
                "ind_transport": _flag(row[5]),
                "ind_devolution": _flag(row[6]),
                "ind_return": _flag(row[7]),
                "ind_annulment": _flag(row[8]),
                "ind_remittance": _flag(row[9]),
                "ind_fuel": _flag(row[10]),
                "ind_excluded_ibs_cbs": _flag(row[11]),
            }
        )
    return {
        "schema": "br.com.planejamento-reforma-tributaria/official-cfop-snapshot",
        "schema_version": "1.0.0",
        "snapshot_id": "CFOP-IT-2023.002-V2.00-2026-08-25",
        "verified_at": "2026-08-31",
        "source": {
            "title": "Tabela de CFOP",
            "technical_version": "IT 2023.002 v2.00",
            "publication_date": "2026-08-25",
            "file_name": "IT_2023.002_v2.00_Tabela_CFOP_indExcI.xlsx",
            "file_url": SOURCE_URL,
            "portal_url": PORTAL_URL,
            "xlsx_sha256": _sha256(source),
        },
        "records": sorted(records, key=lambda item: item["cfop"]),
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-xlsx", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    snapshot = build_snapshot(args.input_xlsx.resolve())
    target = args.output.resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(
        json.dumps(snapshot, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "snapshot_id": snapshot["snapshot_id"],
                "records": len(snapshot["records"]),
                "output": str(target),
            },
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
