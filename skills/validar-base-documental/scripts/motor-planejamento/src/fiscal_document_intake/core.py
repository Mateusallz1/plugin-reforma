from __future__ import annotations

import csv
import hashlib
import json
import re
import unicodedata
from collections import Counter
from collections.abc import Iterable
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from xml.etree.ElementTree import ParseError

from defusedxml import ElementTree as SafeET
from defusedxml.common import DefusedXmlException
from openpyxl import load_workbook
from pypdf import PdfReader
from pypdf.errors import PdfReadError

NFE_NAMESPACE = "http://www.portalfiscal.inf.br/nfe"
CTE_NAMESPACE = "http://www.portalfiscal.inf.br/cte"
MAX_XML_FILES = 5_000
MAX_XML_BYTES = 10 * 1024 * 1024
MAX_PDF_FILES = 5_000
MAX_PDF_BYTES = 50 * 1024 * 1024
MAX_PDF_PAGES = 500
RAW_IGNORED_DIRECTORIES = {
    ".GIT",
    ".VENV",
    "03_SAIDAS",
    "__PYCACHE__",
}
REQUIRED_REPORT_COLUMNS = {
    "document_type",
    "access_key",
    "issue_date",
    "declared_status",
    "gross_amount",
    "source_type",
    "source_name",
    "generated_at",
}
ALLOWED_REPORT_STATUSES = {"AUTHORIZED", "CANCELLED", "UNKNOWN"}
ALLOWED_SOURCE_TYPES = {
    "AUTHORITY_REPORT",
    "ERP_REPORT",
    "ACCOUNTING_REPORT",
    "USER_DECLARED",
}
ANALYSIS_SCOPE_DEFINITIONS = {
    "NFE_NFCE": {
        "label": "NF-e/NFC-e",
        "document_types": ("NFE", "NFCE"),
    },
    "NFSE": {
        "label": "NFS-e",
        "document_types": ("NFSE",),
    },
    "CTE": {
        "label": "CT-e",
        "document_types": ("CTE",),
    },
}
ANALYSIS_GROUP_DEFINITIONS = {
    "NFE_ENTRADAS": {
        "label": "NF-e de entrada",
        "direction": "ENTRADA",
        "document_types": ("NFE",),
    },
    "NFE_SAIDAS": {
        "label": "NF-e de saída",
        "direction": "SAIDA",
        "document_types": ("NFE",),
    },
    "NFCE_ENTRADAS": {
        "label": "NFC-e de entrada",
        "direction": "ENTRADA",
        "document_types": ("NFCE",),
    },
    "NFCE_SAIDAS": {
        "label": "NFC-e de saída",
        "direction": "SAIDA",
        "document_types": ("NFCE",),
    },
    "NFSE_PRESTADOS": {
        "label": "NFS-e de serviços prestados",
        "direction": "SAIDA",
        "document_types": ("NFSE",),
    },
    "NFSE_TOMADOS": {
        "label": "NFS-e de serviços tomados",
        "direction": "ENTRADA",
        "document_types": ("NFSE",),
    },
    "CTE_PRESTADOS": {
        "label": "CT-e de transportes prestados",
        "direction": "SAIDA",
        "document_types": ("CTE",),
    },
    "CTE_TOMADOS": {
        "label": "CT-e de transportes tomados",
        "direction": "ENTRADA",
        "document_types": ("CTE",),
    },
}


class ValidationError(RuntimeError):
    pass


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _digits(value: Any) -> str:
    return re.sub(r"\D", "", str(value or ""))


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _document_ref(source_hash: str) -> str:
    return f"DOC-{source_hash[:12].upper()}"


def _document_ref_for_record(source_hash: str, document_key: str) -> str:
    digest = hashlib.sha256(f"{source_hash}:{document_key}".encode()).hexdigest()
    return f"DOC-{digest[:12].upper()}"


def _report_ref(access_key: str, source_hash: str) -> str:
    digest = hashlib.sha256(f"{access_key}:{source_hash}".encode()).hexdigest()
    return f"RPT-{digest[:12].upper()}"


def validate_access_key(access_key: str) -> bool:
    if len(access_key) != 44 or not access_key.isdigit() or len(set(access_key)) == 1:
        return False
    total = 0
    weight = 2
    for digit in reversed(access_key[:43]):
        total += int(digit) * weight
        weight = 2 if weight == 9 else weight + 1
    remainder = total % 11
    expected = 11 - remainder
    if expected >= 10:
        expected = 0
    return int(access_key[-1]) == expected


def _parse_iso_date(value: str | None) -> date | None:
    if not value:
        return None
    candidate = value.strip().replace("Z", "+00:00")
    try:
        if "T" in candidate:
            return datetime.fromisoformat(candidate).date()
        return date.fromisoformat(candidate[:10])
    except ValueError:
        return None


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    candidate = str(value).strip().replace(" ", "")
    if candidate.count(",") == 1 and candidate.count(".") == 0:
        candidate = candidate.replace(",", ".")
    try:
        parsed = Decimal(candidate)
    except InvalidOperation:
        return None
    return parsed if parsed.is_finite() and parsed >= 0 else None


def _format_decimal(value: Decimal | None) -> str | None:
    return None if value is None else format(value.quantize(Decimal("0.01")), "f")


def _safe_relative_files(folder: Path, pattern: str) -> list[Path]:
    root = folder.resolve()
    paths: list[Path] = []
    for path in sorted(folder.rglob(pattern), key=lambda item: str(item).lower()):
        if path.is_symlink() or not path.is_file():
            continue
        resolved = path.resolve()
        try:
            resolved.relative_to(root)
        except ValueError as error:
            raise ValidationError(
                "Arquivo resolvido fora da pasta autorizada"
            ) from error
        paths.append(resolved)
    return paths


def _direction_hint(path: Path) -> str:
    parts = [part.upper() for part in path.parts]
    if any("SAIDA" in part or "PRESTAD" in part for part in parts):
        return "SAIDA"
    if any("ENTRADA" in part or "TOMAD" in part for part in parts):
        return "ENTRADA"
    return "UNKNOWN"


def _raw_files(folder: Path, suffix: str) -> list[Path]:
    root = folder.resolve()
    paths: list[Path] = []
    for path in sorted(folder.rglob(f"*{suffix}"), key=lambda item: str(item).lower()):
        if path.is_symlink() or not path.is_file():
            continue
        resolved = path.resolve()
        try:
            relative = resolved.relative_to(root)
        except ValueError as error:
            raise ValidationError(
                "Arquivo resolvido fora da pasta autorizada"
            ) from error
        relative_parts = [part.upper() for part in relative.parts[:-1]]
        if any(part.startswith("UC001_") for part in relative_parts):
            continue
        if any(part in RAW_IGNORED_DIRECTORIES for part in relative_parts):
            continue
        paths.append(resolved)
    return paths


def _load_scope(folder: Path) -> dict[str, Any]:
    path = folder / "00_CONTROLE" / "escopo.json"
    if not path.is_file():
        raise ValidationError("Arquivo obrigatório ausente: 00_CONTROLE/escopo.json")
    try:
        scope = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise ValidationError("escopo.json deve ser JSON UTF-8 válido") from error
    if scope.get("schema_version") != "1.0":
        raise ValidationError("schema_version do escopo deve ser 1.0")
    for field in ("entity_ref", "establishment_ref", "period", "analysis_cutoff"):
        if not isinstance(scope.get(field), str) or not scope[field].strip():
            raise ValidationError(f"Campo obrigatório inválido no escopo: {field}")
    if re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", scope["period"]) is None:
        raise ValidationError("period deve usar AAAA-MM")
    if scope.get("objective") != "VALIDATE_DOCUMENT_BASE":
        raise ValidationError("objective deve ser VALIDATE_DOCUMENT_BASE")
    if scope.get("validation_policy") != "DOCUMENTARY_INITIAL":
        raise ValidationError("validation_policy deve ser DOCUMENTARY_INITIAL")
    families = scope.get("document_families")
    if (
        not isinstance(families, list)
        or not families
        or not set(families) <= {"NFE", "NFCE", "NFSE", "CTE"}
    ):
        raise ValidationError(
            "document_families aceita somente NFE, NFCE, NFSE e CTE no UC-001"
        )
    taxpayers = scope.get("entity_taxpayer_ids")
    if not isinstance(taxpayers, list) or not taxpayers:
        raise ValidationError("entity_taxpayer_ids deve conter ao menos um CNPJ")
    normalized = [_digits(value) for value in taxpayers]
    if any(len(value) != 14 for value in normalized):
        raise ValidationError(
            "entity_taxpayer_ids deve conter identificadores numéricos de 14 dígitos"
        )
    scope["entity_taxpayer_ids"] = sorted(set(normalized))
    try:
        datetime.fromisoformat(scope["analysis_cutoff"])
    except ValueError as error:
        raise ValidationError("analysis_cutoff deve usar data/hora ISO") from error
    scope["input_mode"] = "STRUCTURED"
    return scope


def _discover_scope(
    documents: list[dict[str, Any]], xml_paths: list[Path]
) -> dict[str, Any]:
    candidate_counts: Counter[str] = Counter()
    usable_documents = [
        document
        for document in documents
        if document.get("issuer_id") or document.get("recipient_id")
    ]
    for document in usable_documents:
        hint = document.get("folder_direction_hint")
        if hint == "SAIDA" and len(document.get("issuer_id", "")) == 14:
            candidate_counts[document["issuer_id"]] += 1
        elif hint == "ENTRADA" and len(document.get("recipient_id", "")) == 14:
            candidate_counts[document["recipient_id"]] += 1

    if not candidate_counts and usable_documents:
        participation: Counter[str] = Counter()
        for document in usable_documents:
            for taxpayer_id in {
                document.get("issuer_id", ""),
                document.get("recipient_id", ""),
            }:
                if len(taxpayer_id) == 14:
                    participation[taxpayer_id] += 1
        candidate_counts.update(
            {
                taxpayer_id: count
                for taxpayer_id, count in participation.items()
                if count == len(usable_documents)
            }
        )

    if len(candidate_counts) != 1:
        raise ValidationError(
            "Não foi possível identificar uma única empresa na pasta bruta; "
            "organize os documentos em pastas de entrada/saída ou forneça escopo estruturado"
        )
    entity_id = next(iter(candidate_counts))
    periods = sorted(
        {
            document["emission_period"]
            for document in documents
            if document.get("emission_period")
        }
    )
    if len(periods) != 1:
        raise ValidationError(
            "A pasta bruta deve conter uma única competência no UC-001; "
            f"competências encontradas: {', '.join(periods) or 'nenhuma'}"
        )
    families = sorted(
        {
            document["document_type"]
            for document in documents
            if document.get("document_type") in {"NFE", "NFCE", "NFSE", "CTE"}
        }
    )
    if not families:
        raise ValidationError(
            "Nenhuma NF-e/NFC-e/NFS-e/CT-e reconhecida na pasta bruta"
        )
    last_modified = max(path.stat().st_mtime for path in xml_paths)
    entity_hash = hashlib.sha256(entity_id.encode()).hexdigest()[:10].upper()
    return {
        "schema_version": "1.0",
        "entity_ref": f"EMPRESA-{entity_hash}",
        "establishment_ref": f"ESTAB-{entity_hash}",
        "entity_taxpayer_ids": [entity_id],
        "period": periods[0],
        "objective": "VALIDATE_DOCUMENT_BASE",
        "document_families": families,
        "validation_policy": "DOCUMENTARY_INITIAL",
        "analysis_cutoff": datetime.fromtimestamp(last_modified)
        .astimezone()
        .isoformat(timespec="seconds"),
        "input_mode": "RAW_DISCOVERY",
    }


def _normalize_name(value: str) -> str:
    return " ".join(value.split()).strip()


def _company_report_identity(
    documents: list[dict[str, Any]], scope: dict[str, Any]
) -> tuple[dict[str, str], list[dict[str, Any]]]:
    target_ids = set(scope["entity_taxpayer_ids"])
    id_counts: Counter[str] = Counter()
    name_counts: Counter[str] = Counter()
    for document in documents:
        if document.get("issuer_id") in target_ids:
            id_counts[document["issuer_id"]] += 1
            name = _normalize_name(document.get("issuer_name", ""))
            if name:
                name_counts[name] += 1
        if document.get("recipient_id") in target_ids:
            id_counts[document["recipient_id"]] += 1
            name = _normalize_name(document.get("recipient_name", ""))
            if name:
                name_counts[name] += 1

    taxpayer_id = (
        min(id_counts.items(), key=lambda item: (-item[1], item[0]))[0]
        if id_counts
        else min(target_ids)
    )
    company_name = (
        min(name_counts.items(), key=lambda item: (-item[1], item[0]))[0]
        if name_counts
        else "NÃO IDENTIFICADO"
    )
    findings: list[dict[str, Any]] = []
    if not name_counts:
        findings.append({"code": "COMPANY_NAME_NOT_FOUND", "ref": scope["entity_ref"]})
    elif len(name_counts) > 1:
        findings.append({"code": "COMPANY_NAME_VARIATION", "ref": scope["entity_ref"]})
    return {
        "company_name": company_name,
        "company_taxpayer_id": taxpayer_id,
    }, findings


def _format_cnpj(value: str) -> str:
    digits = _digits(value)
    if len(digits) != 14:
        return "NÃO IDENTIFICADO"
    return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"


def _markdown_escape(value: str) -> str:
    escaped = value.replace("\\", "\\\\")
    for character in ("*", "_", "[", "]", "`"):
        escaped = escaped.replace(character, f"\\{character}")
    return escaped


def _find_text(element: Any, path: str) -> str | None:
    found = element.find(path, {"nfe": NFE_NAMESPACE})
    return None if found is None or found.text is None else found.text.strip()


def _parse_cancellation_event(root: Any, source_hash: str) -> dict[str, Any] | None:
    if _local_name(root.tag) != "procEventoNFe":
        return None
    request = root.find("./nfe:evento/nfe:infEvento", {"nfe": NFE_NAMESPACE})
    response = root.find("./nfe:retEvento/nfe:infEvento", {"nfe": NFE_NAMESPACE})
    if request is None or response is None:
        return None
    request_type = _find_text(request, "nfe:tpEvento")
    response_type = _find_text(response, "nfe:tpEvento")
    request_key = _digits(_find_text(request, "nfe:chNFe"))
    response_key = _digits(_find_text(response, "nfe:chNFe"))
    request_sequence = _find_text(request, "nfe:nSeqEvento")
    response_sequence = _find_text(response, "nfe:nSeqEvento")
    response_status = _find_text(response, "nfe:cStat")
    protocol = _find_text(response, "nfe:nProt")
    confirmed = (
        request_type == response_type == "110111"
        and request_key == response_key
        and validate_access_key(request_key)
        and request_sequence == response_sequence
        and response_status in {"135", "155"}
        and bool(protocol and protocol.isdigit())
    )
    return {
        "access_key": request_key,
        "event_type": "CANCELLATION",
        "confirmed": confirmed,
        "source_hash": source_hash,
    }


def _parse_document(root: Any, source_hash: str) -> dict[str, Any] | None:
    root_name = _local_name(root.tag)
    if root_name == "nfeProc":
        nfe = root.find("./nfe:NFe", {"nfe": NFE_NAMESPACE})
        protocol = root.find("./nfe:protNFe/nfe:infProt", {"nfe": NFE_NAMESPACE})
    elif root_name == "NFe":
        nfe = root
        protocol = None
    else:
        return None
    if nfe is None:
        return None
    info = nfe.find("./nfe:infNFe", {"nfe": NFE_NAMESPACE})
    if info is None:
        return {
            "source_hash": source_hash,
            "document_ref": _document_ref(source_hash),
            "document_key": f"INVALID:{source_hash}",
            "access_key": "",
            "document_type": "UNKNOWN",
            "analysis_scope": "NFE_NFCE",
            "validity_status": "INVALID_STRUCTURE",
            "reason_codes": ["INFNFE_MISSING"],
            "warnings": [],
            "gross_amount": None,
            "item_count": 0,
            "emission_period": None,
            "issuer_id": "",
            "issuer_name": "",
            "recipient_id": "",
            "recipient_name": "",
        }

    access_key = _digits(info.attrib.get("Id", "").removeprefix("NFe"))
    model = _find_text(info, "nfe:ide/nfe:mod")
    emission_raw = _find_text(info, "nfe:ide/nfe:dhEmi") or _find_text(
        info, "nfe:ide/nfe:dEmi"
    )
    emission = _parse_iso_date(emission_raw)
    issuer_id = _digits(
        _find_text(info, "nfe:emit/nfe:CNPJ") or _find_text(info, "nfe:emit/nfe:CPF")
    )
    issuer_name = _find_text(info, "nfe:emit/nfe:xNome") or _find_text(
        info, "nfe:emit/nfe:xFant"
    )
    recipient_id = _digits(
        _find_text(info, "nfe:dest/nfe:CNPJ") or _find_text(info, "nfe:dest/nfe:CPF")
    )
    recipient_name = _find_text(info, "nfe:dest/nfe:xNome") or ""
    gross_amount = _parse_decimal(_find_text(info, "nfe:total/nfe:ICMSTot/nfe:vNF"))
    item_count = len(info.findall("./nfe:det", {"nfe": NFE_NAMESPACE}))
    document_type = {"55": "NFE", "65": "NFCE"}.get(model, "UNKNOWN")

    reasons: list[str] = []
    if not validate_access_key(access_key):
        reasons.append("ACCESS_KEY_INVALID")
    if document_type == "UNKNOWN":
        reasons.append("MODEL_UNSUPPORTED")
    if emission is None:
        reasons.append("EMISSION_DATE_INVALID")
    if gross_amount is None:
        reasons.append("GROSS_AMOUNT_INVALID")
    if len(issuer_id) not in {11, 14}:
        reasons.append("ISSUER_ID_INVALID")
    if access_key and len(issuer_id) == 14 and access_key[6:20] != issuer_id:
        reasons.append("ACCESS_KEY_ISSUER_MISMATCH")

    protocol_ok = False
    if protocol is not None:
        protocol_key = _digits(_find_text(protocol, "nfe:chNFe"))
        status = _find_text(protocol, "nfe:cStat")
        protocol_number = _find_text(protocol, "nfe:nProt")
        protocol_ok = (
            status == "100"
            and protocol_key == access_key
            and bool(protocol_number and protocol_number.isdigit())
        )
        if not protocol_ok:
            reasons.append("AUTHORIZATION_PROTOCOL_INVALID")

    if reasons:
        validity = "INVALID_STRUCTURE"
    elif protocol_ok:
        validity = "VALID_DOCUMENTARY"
    else:
        validity = "STATUS_NOT_VERIFIABLE"

    warnings = [
        "AUTHORITY_CURRENT_STATUS_NOT_CONFIRMED",
        "SIGNATURE_NOT_VERIFIED",
        "XSD_NOT_VALIDATED",
    ]
    return {
        "source_hash": source_hash,
        "document_ref": _document_ref(source_hash),
        "document_key": access_key or f"NFE:{source_hash}",
        "access_key": access_key,
        "document_type": document_type,
        "analysis_scope": "NFE_NFCE",
        "validity_status": validity,
        "reason_codes": reasons,
        "warnings": warnings,
        "gross_amount": gross_amount,
        "item_count": item_count,
        "emission_period": emission.strftime("%Y-%m") if emission else None,
        "issuer_id": issuer_id,
        "issuer_name": issuer_name or "",
        "recipient_id": recipient_id,
        "recipient_name": recipient_name,
        "pdf_match_tokens": [access_key] if access_key else [],
    }


def _find_cte_text(element: Any, path: str) -> str | None:
    found = element.find(path, {"cte": CTE_NAMESPACE})
    return None if found is None or found.text is None else found.text.strip()


def _parse_cte_document(root: Any, source_hash: str) -> dict[str, Any] | None:
    root_name = _local_name(root.tag)
    if root_name == "cteProc":
        cte = root.find("./cte:CTe", {"cte": CTE_NAMESPACE})
        protocol = root.find("./cte:protCTe/cte:infProt", {"cte": CTE_NAMESPACE})
    elif root_name == "CTe":
        cte = root
        protocol = None
    else:
        return None
    if cte is None:
        return None
    info = cte.find("./cte:infCte", {"cte": CTE_NAMESPACE})
    if info is None:
        return {
            "source_hash": source_hash,
            "document_ref": _document_ref(source_hash),
            "document_key": f"INVALID-CTE:{source_hash}",
            "access_key": "",
            "document_type": "UNKNOWN",
            "analysis_scope": "CTE",
            "validity_status": "INVALID_STRUCTURE",
            "reason_codes": ["INFCTE_MISSING"],
            "warnings": [],
            "gross_amount": None,
            "item_count": 0,
            "emission_period": None,
            "issuer_id": "",
            "issuer_name": "",
            "recipient_id": "",
            "recipient_name": "",
            "taker_role": "NOT_IDENTIFIED",
            "pdf_match_tokens": [],
        }

    access_key = _digits(info.attrib.get("Id", "").removeprefix("CTe"))
    model = _find_cte_text(info, "cte:ide/cte:mod")
    emission_raw = _find_cte_text(info, "cte:ide/cte:dhEmi") or _find_cte_text(
        info, "cte:ide/cte:dEmi"
    )
    emission = _parse_iso_date(emission_raw)
    issuer_id = _digits(
        _find_cte_text(info, "cte:emit/cte:CNPJ")
        or _find_cte_text(info, "cte:emit/cte:CPF")
    )
    issuer_name = _find_cte_text(info, "cte:emit/cte:xNome") or ""
    gross_amount = _parse_decimal(_find_cte_text(info, "cte:vPrest/cte:vTPrest"))

    party_paths = {
        "SENDER": "cte:rem",
        "DISPATCHER": "cte:exped",
        "RECEIVER": "cte:receb",
        "DESTINATION": "cte:dest",
    }
    party_ids: dict[str, str] = {}
    party_names: dict[str, str] = {}
    for role, path in party_paths.items():
        party_ids[role] = _digits(
            _find_cte_text(info, f"{path}/cte:CNPJ")
            or _find_cte_text(info, f"{path}/cte:CPF")
        )
        party_names[role] = _find_cte_text(info, f"{path}/cte:xNome") or ""

    taker_code = _find_cte_text(info, "cte:ide/cte:toma3/cte:toma")
    taker_role = {
        "0": "SENDER",
        "1": "DISPATCHER",
        "2": "RECEIVER",
        "3": "DESTINATION",
    }.get(taker_code or "", "OTHER" if taker_code == "4" else "NOT_IDENTIFIED")
    if taker_role in party_ids:
        taker_id = party_ids[taker_role]
        taker_name = party_names[taker_role]
    else:
        taker_id = _digits(
            _find_cte_text(info, "cte:ide/cte:toma4/cte:CNPJ")
            or _find_cte_text(info, "cte:ide/cte:toma4/cte:CPF")
        )
        taker_name = _find_cte_text(info, "cte:ide/cte:toma4/cte:xNome") or ""

    reasons: list[str] = []
    if not validate_access_key(access_key):
        reasons.append("ACCESS_KEY_INVALID")
    if model != "57":
        reasons.append("CTE_MODEL_UNSUPPORTED")
    if emission is None:
        reasons.append("EMISSION_DATE_INVALID")
    if gross_amount is None:
        reasons.append("GROSS_AMOUNT_INVALID")
    if len(issuer_id) != 14:
        reasons.append("ISSUER_ID_INVALID")
    if access_key and len(issuer_id) == 14 and access_key[6:20] != issuer_id:
        reasons.append("ACCESS_KEY_ISSUER_MISMATCH")
    if len(taker_id) not in {11, 14}:
        reasons.append("CTE_TAKER_NOT_IDENTIFIED")

    protocol_ok = False
    if protocol is not None:
        protocol_key = _digits(_find_cte_text(protocol, "cte:chCTe"))
        status = _find_cte_text(protocol, "cte:cStat")
        protocol_number = _find_cte_text(protocol, "cte:nProt")
        protocol_ok = (
            status == "100"
            and protocol_key == access_key
            and bool(protocol_number and protocol_number.isdigit())
        )
        if not protocol_ok:
            reasons.append("AUTHORIZATION_PROTOCOL_INVALID")

    if reasons:
        validity = "INVALID_STRUCTURE"
    elif protocol_ok:
        validity = "VALID_DOCUMENTARY"
    else:
        validity = "STATUS_NOT_VERIFIABLE"

    return {
        "source_hash": source_hash,
        "document_ref": _document_ref(source_hash),
        "document_key": access_key or f"CTE:{source_hash}",
        "access_key": access_key,
        "document_type": "CTE" if model == "57" else "UNKNOWN",
        "analysis_scope": "CTE",
        "validity_status": validity,
        "reason_codes": reasons,
        "warnings": [
            "AUTHORITY_CURRENT_STATUS_NOT_CONFIRMED",
            "SIGNATURE_NOT_VERIFIED",
            "XSD_NOT_VALIDATED",
        ],
        "gross_amount": gross_amount,
        "item_count": 1,
        "emission_period": emission.strftime("%Y-%m") if emission else None,
        "issuer_id": issuer_id,
        "issuer_name": issuer_name,
        "recipient_id": taker_id,
        "recipient_name": taker_name,
        "taker_role": taker_role,
        "pdf_match_tokens": [access_key] if access_key else [],
    }


def _find_local_element(element: Any, names: set[str]) -> Any | None:
    normalized = {name.lower() for name in names}
    return next(
        (
            child
            for child in element.iter()
            if _local_name(child.tag).lower() in normalized
        ),
        None,
    )


def _find_local_text(element: Any | None, names: set[str]) -> str | None:
    if element is None:
        return None
    found = _find_local_element(element, names)
    if found is None or found.text is None:
        return None
    value = found.text.strip()
    return value or None


def _nfse_status(value: str | None) -> str:
    if not value:
        return "NOT_EMBEDDED"
    normalized = re.sub(r"[^A-Z0-9]", "", value.upper())
    if normalized in {"2", "CANCELADA", "CANCELADO"} or "CANCEL" in normalized:
        return "CANCELLED"
    if normalized in {"1", "ATIVA", "ATIVO", "EMITIDA", "EMITIDO", "NORMAL"}:
        return "ACTIVE"
    return "UNKNOWN"


def _parse_nfse_documents(root: Any, source_hash: str) -> list[dict[str, Any]] | None:
    notes = [
        element
        for element in root.iter()
        if _local_name(element.tag).lower() == "infnfse"
    ]
    if not notes:
        return None

    documents: list[dict[str, Any]] = []
    for index, note in enumerate(notes, start=1):
        provider = _find_local_element(
            note, {"PrestadorServico", "IdentificacaoPrestador", "Prestador"}
        )
        recipient = _find_local_element(
            note, {"TomadorServico", "IdentificacaoTomador", "Tomador"}
        )
        service = _find_local_element(note, {"Servico"})
        number = _find_local_text(note, {"Numero"}) or ""
        verification_code = _find_local_text(note, {"CodigoVerificacao"}) or ""
        national_access_key = _digits(
            _find_local_text(
                note,
                {
                    "ChaveAcesso",
                    "ChaveAcessoNfse",
                    "ChaveAcessoNFSe",
                },
            )
        )
        emission_raw = _find_local_text(note, {"DataEmissao"}) or _find_local_text(
            note, {"Competencia"}
        )
        emission = _parse_iso_date(emission_raw)
        issuer_id = _digits(_find_local_text(provider, {"Cnpj", "Cpf"}))
        recipient_id = _digits(_find_local_text(recipient, {"Cnpj", "Cpf"}))
        issuer_name = _find_local_text(provider, {"RazaoSocial", "NomeFantasia"}) or ""
        recipient_name = (
            _find_local_text(recipient, {"RazaoSocial", "NomeFantasia"}) or ""
        )
        service_context = service if service is not None else note
        gross_amount = _parse_decimal(
            _find_local_text(service_context, {"ValorServicos", "vServ", "vServicos"})
        )
        municipality_code = _digits(
            _find_local_text(service_context, {"CodigoMunicipio"})
        )
        status = _nfse_status(
            _find_local_text(note, {"Situacao", "StatusNfse", "StatusNFSe"})
        )

        reasons: list[str] = []
        if not number:
            reasons.append("NFSE_NUMBER_MISSING")
        if not verification_code and not national_access_key:
            reasons.append("NFSE_VERIFICATION_IDENTIFIER_MISSING")
        if emission is None:
            reasons.append("EMISSION_DATE_INVALID")
        if gross_amount is None:
            reasons.append("GROSS_AMOUNT_INVALID")
        if len(issuer_id) not in {11, 14}:
            reasons.append("ISSUER_ID_INVALID")
        if len(recipient_id) not in {11, 14}:
            reasons.append("RECIPIENT_ID_INVALID")
        if status == "UNKNOWN":
            reasons.append("NFSE_STATUS_UNKNOWN")

        canonical_material = (
            f"NFSE|{issuer_id}|{municipality_code}|{number}|"
            f"{verification_code.upper()}|{national_access_key}"
        )
        if not any(
            [
                issuer_id,
                municipality_code,
                number,
                verification_code,
                national_access_key,
            ]
        ):
            canonical_material = f"NFSE|{source_hash}|{index}"
        document_key = "NFSE-" + hashlib.sha256(canonical_material.encode()).hexdigest()
        pdf_tokens = [
            token
            for token in {verification_code, national_access_key}
            if len(re.sub(r"[^0-9A-Za-z]", "", token)) >= 6
        ]

        if status == "CANCELLED":
            validity = "CANCELLED"
            reasons.append("NFSE_CANCELLED_IN_DOCUMENT")
        elif reasons:
            validity = "INVALID_STRUCTURE"
        else:
            validity = "VALID_DOCUMENTARY"

        warnings = [
            "AUTHORITY_CURRENT_STATUS_NOT_CONFIRMED",
            "SIGNATURE_NOT_VERIFIED",
            "XSD_NOT_VALIDATED",
        ]
        if status == "NOT_EMBEDDED":
            warnings.append("NFSE_STATUS_NOT_EMBEDDED")

        documents.append(
            {
                "source_hash": source_hash,
                "document_ref": _document_ref_for_record(source_hash, document_key),
                "document_key": document_key,
                "access_key": national_access_key,
                "document_type": "NFSE",
                "analysis_scope": "NFSE",
                "validity_status": validity,
                "reason_codes": reasons,
                "warnings": warnings,
                "gross_amount": gross_amount,
                "item_count": 1 if service is not None else 0,
                "emission_period": emission.strftime("%Y-%m") if emission else None,
                "issuer_id": issuer_id,
                "issuer_name": issuer_name,
                "recipient_id": recipient_id,
                "recipient_name": recipient_name,
                "pdf_match_tokens": sorted(pdf_tokens),
                "pdf_report_tokens": [number] if len(_digits(number)) >= 3 else [],
            }
        )
    return documents


def _parse_xml_file(
    path: Path,
) -> tuple[list[dict[str, Any]], dict[str, Any] | None, dict[str, Any] | None]:
    source_hash = _sha256_file(path)
    if path.stat().st_size > MAX_XML_BYTES:
        return (
            [],
            None,
            {
                "document_ref": _document_ref(source_hash),
                "source_hash": source_hash,
                "status": "INVALID_STRUCTURE",
                "analysis_scope": "UNKNOWN",
                "reason_codes": ["XML_FILE_TOO_LARGE"],
            },
        )
    try:
        root = SafeET.parse(path).getroot()
    except (OSError, ParseError, DefusedXmlException):
        return (
            [],
            None,
            {
                "document_ref": _document_ref(source_hash),
                "source_hash": source_hash,
                "status": "INVALID_STRUCTURE",
                "analysis_scope": "UNKNOWN",
                "reason_codes": ["XML_NOT_PARSEABLE"],
            },
        )
    event = _parse_cancellation_event(root, source_hash)
    if event is not None:
        return [], event, None
    document = _parse_document(root, source_hash)
    if document is not None:
        return [document], None, None
    cte_document = _parse_cte_document(root, source_hash)
    if cte_document is not None:
        return [cte_document], None, None
    nfse_documents = _parse_nfse_documents(root, source_hash)
    if nfse_documents is not None:
        return nfse_documents, None, None
    return (
        [],
        None,
        {
            "document_ref": _document_ref(source_hash),
            "source_hash": source_hash,
            "status": "UNSUPPORTED_LAYOUT",
            "analysis_scope": "UNKNOWN",
            "reason_codes": [f"ROOT_{_local_name(root.tag).upper()}_UNSUPPORTED"],
        },
    )


def _candidate_access_keys(text: str) -> set[str]:
    candidates = set(re.findall(r"(?<!\d)\d{44}(?!\d)", text))
    for match in re.finditer(r"(?:\d[\s.\-]*){44}", text):
        digits = _digits(match.group(0))
        if len(digits) == 44:
            candidates.add(digits)
    return {candidate for candidate in candidates if validate_access_key(candidate)}


def _fold_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(
        character for character in normalized if not unicodedata.combining(character)
    ).lower()


def _compact_search_text(value: str) -> str:
    return re.sub(r"[^0-9a-z]", "", _fold_text(value))


def _parse_pdf_file(
    path: Path, documents_by_key: dict[str, dict[str, Any]]
) -> dict[str, Any]:
    source_hash = _sha256_file(path)
    pdf_ref = f"PDF-{source_hash[:12].upper()}"
    base_record = {
        "pdf_ref": pdf_ref,
        "source_hash": source_hash,
        "direction_hint": _direction_hint(path),
        "page_count": 0,
        "keys_detected": 0,
        "identifiers_detected": 0,
        "matched_document_refs": [],
        "unmatched_keys": 0,
        "unmatched_identifiers": 0,
        "pdf_kind": "UNKNOWN",
        "status": "PDF_READ_ERROR",
        "warnings": [],
    }
    if path.stat().st_size > MAX_PDF_BYTES:
        base_record["warnings"].append("PDF_FILE_TOO_LARGE")
        return base_record
    try:
        reader = PdfReader(path)
        if reader.is_encrypted and reader.decrypt("") == 0:
            base_record["warnings"].append("PDF_ENCRYPTED")
            return base_record
        page_count = len(reader.pages)
        base_record["page_count"] = page_count
        if page_count == 0 or page_count > MAX_PDF_PAGES:
            base_record["warnings"].append("PDF_PAGE_LIMIT_INVALID")
            return base_record
        extracted_text = "\n".join((page.extract_text() or "") for page in reader.pages)
    except (OSError, PdfReadError, ValueError, TypeError, KeyError):
        base_record["warnings"].append("PDF_NOT_READABLE")
        return base_record

    raw_keys = _candidate_access_keys(path.stem) | _candidate_access_keys(
        extracted_text
    )
    documents_by_access_key = {
        document["access_key"]: document
        for document in documents_by_key.values()
        if document["document_type"] in {"NFE", "NFCE", "CTE"}
        and document.get("access_key")
    }
    folded_text = _fold_text(extracted_text)
    compact_text = _compact_search_text(f"{path.stem}\n{extracted_text}")
    looks_like_nfse = "nfse" in _compact_search_text(extracted_text) or (
        "nota fiscal" in folded_text and "servi" in folded_text
    )
    looks_like_danfe = "danfe" in folded_text or (
        "documento auxiliar" in folded_text and "nota fiscal eletronica" in folded_text
    )
    looks_like_dacte = "dacte" in folded_text or (
        "documento auxiliar" in folded_text
        and "conhecimento de transporte" in folded_text
    )
    is_nfse_report = (
        "termo de abertura" in folded_text
        or "registros de notas fiscais" in folded_text
    )
    is_nfse_print = "dados da nfse" in folded_text or (
        "codigo de verificacao" in folded_text and "tomador do servico" in folded_text
    )
    keys = (
        raw_keys & set(documents_by_access_key)
        if looks_like_nfse and not looks_like_danfe and not looks_like_dacte
        else raw_keys
    )
    if looks_like_dacte:
        eligible_access_types = {"CTE"}
    elif looks_like_danfe:
        eligible_access_types = {"NFE", "NFCE"}
    else:
        eligible_access_types = {"NFE", "NFCE", "CTE"}
    eligible_by_access_key = {
        key: document
        for key, document in documents_by_access_key.items()
        if document["document_type"] in eligible_access_types
    }
    matched_by_ref = {
        eligible_by_access_key[key]["document_ref"]: eligible_by_access_key[key]
        for key in sorted(keys & set(eligible_by_access_key))
    }
    hint = base_record["direction_hint"]
    for document in documents_by_key.values():
        if document["document_type"] != "NFSE":
            continue
        if hint in {"ENTRADA", "SAIDA"} and document.get("direction") != hint:
            continue
        tokens = (
            document.get("pdf_report_tokens", [])
            if is_nfse_report and not is_nfse_print
            else document.get("pdf_match_tokens", [])
        )
        if is_nfse_report and not is_nfse_print:
            token_matches = any(
                _compact_search_text(token) in compact_text
                for token in tokens
                if _compact_search_text(token)
            )
        else:
            token_matches = any(
                _compact_search_text(token) in compact_text
                for token in tokens
                if _compact_search_text(token)
            )
        if token_matches:
            matched_by_ref[document["document_ref"]] = document
    matched = list(matched_by_ref.values())
    unmatched = keys - set(eligible_by_access_key)
    matched_nfse = [
        document for document in matched if document["document_type"] == "NFSE"
    ]
    matched_cte = [
        document for document in matched if document["document_type"] == "CTE"
    ]
    base_record["keys_detected"] = len(keys)
    base_record["identifiers_detected"] = len(keys) + len(matched_nfse)
    base_record["matched_document_refs"] = sorted(
        {document["document_ref"] for document in matched}
    )
    base_record["unmatched_keys"] = len(unmatched)
    base_record["unmatched_identifiers"] = len(unmatched)

    matched_types = {document["document_type"] for document in matched}
    if looks_like_dacte:
        base_record["pdf_kind"] = "DACTE"
        if matched_cte:
            base_record["status"] = "DACTE_MATCHED"
        else:
            base_record["status"] = "DACTE_WITHOUT_XML"
            base_record["warnings"].append("DACTE_WITHOUT_XML")
    elif not keys and not matched_nfse and looks_like_nfse:
        base_record["pdf_kind"] = (
            "NFSE_REPORT" if is_nfse_report and not is_nfse_print else "NFSE_PRINT"
        )
        base_record["status"] = "NFSE_PDF_WITHOUT_XML"
        base_record["warnings"].append("NFSE_PDF_WITHOUT_XML")
    elif not keys and not matched_nfse:
        base_record["status"] = "PDF_KEY_NOT_FOUND"
        base_record["warnings"].append("PDF_KEY_NOT_FOUND")
    elif unmatched:
        base_record["status"] = "DANFE_WITHOUT_XML"
        base_record["warnings"].append("DANFE_WITHOUT_XML")
    elif matched_types == {"NFSE"}:
        if is_nfse_report and not is_nfse_print:
            base_record["pdf_kind"] = "NFSE_REPORT"
            base_record["status"] = "NFSE_REPORT_MATCHED"
        else:
            base_record["pdf_kind"] = "NFSE_PRINT"
            base_record["status"] = "NFSE_PDF_MATCHED"
    elif "NFSE" in matched_types:
        base_record["pdf_kind"] = "MIXED_FISCAL_PDF"
        base_record["status"] = "FISCAL_PDF_MATCHED"
    else:
        base_record["pdf_kind"] = "DANFE"
        base_record["status"] = "DANFE_MATCHED"

    matched_directions = {
        document.get("direction")
        for document in matched
        if document.get("direction") in {"ENTRADA", "SAIDA"}
    }
    if hint in {"ENTRADA", "SAIDA"} and any(
        direction != hint for direction in matched_directions
    ):
        base_record["warnings"].append(
            "DACTE_DIRECTION_CONFLICT"
            if looks_like_dacte
            else "NFSE_PDF_DIRECTION_CONFLICT"
            if matched_types == {"NFSE"}
            else "DANFE_DIRECTION_CONFLICT"
        )
    return base_record


def _load_csv_report(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as source:
        return [dict(row) for row in csv.DictReader(source)]


def _load_xlsx_report(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = next(
            sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"
        )
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        return [
            dict(zip(headers, row, strict=False))
            for row in rows
            if any(value is not None for value in row)
        ]
    except (StopIteration, ValueError) as error:
        raise ValidationError(
            f"Relatório XLSX sem planilha ou cabeçalho: {path.name}"
        ) from error
    finally:
        workbook.close()


def _load_reports(folder: Path) -> tuple[list[dict[str, Any]], list[str]]:
    report_dir = folder / "02_RELATORIOS"
    if not report_dir.is_dir():
        return [], []
    report_paths = [
        path
        for path in _safe_relative_files(report_dir, "*")
        if path.suffix.lower() in {".csv", ".xlsx"}
    ]
    records: list[dict[str, Any]] = []
    hashes: list[str] = []
    for path in report_paths:
        hashes.append(_sha256_file(path))
        rows = (
            _load_csv_report(path)
            if path.suffix.lower() == ".csv"
            else _load_xlsx_report(path)
        )
        for index, row in enumerate(rows, start=2):
            missing = REQUIRED_REPORT_COLUMNS - set(row)
            if missing:
                raise ValidationError(
                    f"Relatório {path.name} sem colunas: {', '.join(sorted(missing))}"
                )
            access_key = _digits(row.get("access_key"))
            gross_amount = _parse_decimal(row.get("gross_amount"))
            status = str(row.get("declared_status") or "").strip().upper()
            source_type = str(row.get("source_type") or "").strip().upper()
            document_type = str(row.get("document_type") or "").strip().upper()
            issue_date = _parse_iso_date(str(row.get("issue_date") or ""))
            generated_at = str(row.get("generated_at") or "").strip()
            errors: list[str] = []
            if not validate_access_key(access_key):
                errors.append("REPORT_ACCESS_KEY_INVALID")
            if gross_amount is None:
                errors.append("REPORT_GROSS_AMOUNT_INVALID")
            if status not in ALLOWED_REPORT_STATUSES:
                errors.append("REPORT_STATUS_INVALID")
            if source_type not in ALLOWED_SOURCE_TYPES:
                errors.append("REPORT_SOURCE_TYPE_INVALID")
            if document_type not in {"NFE", "NFCE", "CTE"}:
                errors.append("REPORT_DOCUMENT_TYPE_INVALID")
            if issue_date is None:
                errors.append("REPORT_ISSUE_DATE_INVALID")
            try:
                datetime.fromisoformat(generated_at)
            except ValueError:
                errors.append("REPORT_GENERATED_AT_INVALID")
            records.append(
                {
                    "report_ref": _report_ref(access_key, hashes[-1]),
                    "document_key": access_key,
                    "access_key": access_key,
                    "document_type": document_type,
                    "issue_period": issue_date.strftime("%Y-%m")
                    if issue_date
                    else None,
                    "declared_status": status,
                    "gross_amount": gross_amount,
                    "source_type": source_type,
                    "source_name": str(row.get("source_name") or "").strip(),
                    "errors": errors,
                    "row": index,
                }
            )
    return records, hashes


def _apply_document_rules(
    documents: list[dict[str, Any]],
    events: list[dict[str, Any]],
    scope: dict[str, Any],
) -> None:
    cancelled_keys = {
        event["access_key"]
        for event in events
        if event.get("confirmed") and event.get("access_key")
    }
    seen_keys: set[str] = set()
    target_ids = set(scope["entity_taxpayer_ids"])
    for document in documents:
        access_key = document["access_key"]
        document_key = document["document_key"]
        if access_key and access_key in cancelled_keys:
            document["validity_status"] = "CANCELLED"
            document["reason_codes"].append("CONFIRMED_CANCELLATION_EVENT")
        if (
            document["emission_period"]
            and document["emission_period"] != scope["period"]
        ):
            document["scope_status"] = "OUT_OF_PERIOD"
            document["validity_status"] = "OUT_OF_PERIOD"
            document["reason_codes"].append("EMISSION_OUT_OF_PERIOD")
        elif not ({document["issuer_id"], document["recipient_id"]} & target_ids):
            document["scope_status"] = "OUT_OF_SCOPE"
            document["validity_status"] = "OUT_OF_SCOPE"
            document["reason_codes"].append("TARGET_NOT_PARTY")
        else:
            document["scope_status"] = "IN_SCOPE"
            if (
                document["issuer_id"] in target_ids
                and document["recipient_id"] in target_ids
            ):
                document["direction"] = "BOTH"
            elif document["issuer_id"] in target_ids:
                document["direction"] = "SAIDA"
            else:
                document["direction"] = "ENTRADA"
            folder_hint = document.get("folder_direction_hint")
            if (
                folder_hint in {"ENTRADA", "SAIDA"}
                and folder_hint != document["direction"]
            ):
                document["warnings"].append("FOLDER_DIRECTION_CONFLICT")
        if document_key in seen_keys:
            document["validity_status"] = "DUPLICATE"
            document["reason_codes"].append(
                "ACCESS_KEY_ALREADY_SEEN"
                if document["document_type"] in {"NFE", "NFCE", "CTE"}
                else "NFSE_DOCUMENT_ALREADY_SEEN"
            )
        else:
            seen_keys.add(document_key)
        document.setdefault("direction", "NAO_VERIFICAVEL")
        document["included"] = (
            document["validity_status"] == "VALID_DOCUMENTARY"
            and document["scope_status"] == "IN_SCOPE"
        )


def _reconcile(
    documents: list[dict[str, Any]],
    reports: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    primary_by_key: dict[str, dict[str, Any]] = {}
    for document in documents:
        key = document["document_key"]
        if (
            key
            and document["validity_status"] != "DUPLICATE"
            and key not in primary_by_key
        ):
            primary_by_key[key] = document

    reconciliations: list[dict[str, Any]] = []
    report_keys: set[str] = set()
    for report in reports:
        key = report["document_key"]
        if key in report_keys:
            status = "REPORT_DUPLICATE"
            document = primary_by_key.get(key)
        else:
            report_keys.add(key)
            document = primary_by_key.get(key)
            if report["errors"]:
                status = "REPORT_INVALID"
            elif document is None:
                status = "DECLARED_WITHOUT_XML"
            elif document["validity_status"] in {
                "INVALID_STRUCTURE",
                "STATUS_NOT_VERIFIABLE",
            }:
                status = "STRUCTURALLY_UNAVAILABLE_WITH_REPORT"
            elif (
                report["declared_status"] == "AUTHORIZED"
                and document["validity_status"] == "CANCELLED"
            ) or (
                report["declared_status"] == "CANCELLED"
                and document["validity_status"] == "VALID_DOCUMENTARY"
            ):
                status = "STATUS_CONFLICT"
            elif (
                report["declared_status"] == "CANCELLED"
                and document["validity_status"] == "CANCELLED"
            ):
                status = "MATCHED_CANCELLED"
            elif report["document_type"] != document["document_type"]:
                status = "TYPE_MISMATCH"
            elif report["issue_period"] != document["emission_period"]:
                status = "PERIOD_MISMATCH"
            elif (
                document["gross_amount"] is not None
                and report["gross_amount"] is not None
                and document["gross_amount"] != report["gross_amount"]
            ):
                status = "VALUE_MISMATCH"
            elif document["validity_status"] in {
                "OUT_OF_PERIOD",
                "OUT_OF_SCOPE",
                "UNSUPPORTED_LAYOUT",
            }:
                status = "DOCUMENT_NOT_ELIGIBLE"
            else:
                status = "MATCHED_VALID"
        reconciliations.append(
            {
                "report_ref": report["report_ref"],
                "document_ref": document["document_ref"] if document else None,
                "status": status,
                "reported_amount": _format_decimal(report["gross_amount"]),
                "documented_amount": _format_decimal(document["gross_amount"])
                if document
                else None,
                "source_type": report["source_type"],
                "reason_codes": report["errors"],
            }
        )

    for key, document in primary_by_key.items():
        if key not in report_keys and document["scope_status"] == "IN_SCOPE":
            reconciliations.append(
                {
                    "report_ref": None,
                    "document_ref": document["document_ref"],
                    "status": "XML_WITHOUT_REPORT",
                    "reported_amount": None,
                    "documented_amount": _format_decimal(document["gross_amount"]),
                    "source_type": None,
                    "reason_codes": [],
                }
            )
    return reconciliations


def _sum_decimal(values: Iterable[Decimal | None]) -> Decimal:
    return sum((value for value in values if value is not None), Decimal(0))


def _analysis_group(document: dict[str, Any]) -> str:
    for code, definition in ANALYSIS_GROUP_DEFINITIONS.items():
        if (
            document.get("direction") == definition["direction"]
            and document.get("document_type") in definition["document_types"]
        ):
            return code
    return "NAO_CLASSIFICADO"


def _analysis_scope(document: dict[str, Any]) -> str:
    explicit = document.get("analysis_scope")
    if explicit:
        return explicit
    for code, definition in ANALYSIS_SCOPE_DEFINITIONS.items():
        if document.get("document_type") in definition["document_types"]:
            return code
    return "UNKNOWN"


def _build_scope_authorizations(
    documents: list[dict[str, Any]],
    file_errors: list[dict[str, Any]],
    blockers: list[dict[str, Any]],
    declared_families: list[str],
) -> tuple[dict[str, Any], list[str], list[str]]:
    declared = set(declared_families)
    authorizations: dict[str, Any] = {}
    authorized_scopes: list[str] = []
    restricted_scopes: list[str] = []
    for code, definition in ANALYSIS_SCOPE_DEFINITIONS.items():
        scope_documents = [
            document for document in documents if _analysis_scope(document) == code
        ]
        scope_errors = [
            error for error in file_errors if error.get("analysis_scope") == code
        ]
        declared_scope = bool(declared & set(definition["document_types"]))
        detected = bool(scope_documents or scope_errors)
        included = [document for document in scope_documents if document["included"]]
        blocker_codes = sorted(
            {item["code"] for item in blockers if item.get("scope") == code}
        )
        if (declared_scope or detected) and not included and not blocker_codes:
            blocker_codes.append("NO_ELIGIBLE_DOCUMENTS_IN_SCOPE")
        authorized = bool(included) and not blocker_codes
        if authorized:
            status = "READY"
            authorized_scopes.append(code)
        elif declared_scope or detected:
            status = "BLOCKED"
            restricted_scopes.append(code)
        else:
            status = "NOT_DETECTED"
        authorizations[code] = {
            "label": definition["label"],
            "declared": declared_scope,
            "detected": detected,
            "status": status,
            "authorized": authorized,
            "document_types": list(definition["document_types"]),
            "included": len(included),
            "excluded": len(scope_documents) - len(included) + len(scope_errors),
            "blocker_codes": blocker_codes,
        }

    if any(item.get("scope") == "UNKNOWN" for item in blockers):
        restricted_scopes.append("UNKNOWN")
    return authorizations, authorized_scopes, sorted(set(restricted_scopes))


def validate_folder(folder: Path | str) -> dict[str, Any]:
    base = Path(folder).expanduser().resolve()
    if not base.is_dir():
        raise ValidationError("A pasta informada não existe")
    structured_input = (base / "00_CONTROLE" / "escopo.json").is_file() and (
        base / "01_XML"
    ).is_dir()
    if structured_input:
        scope: dict[str, Any] | None = _load_scope(base)
        xml_dir = base / "01_XML"
        xml_paths = _safe_relative_files(xml_dir, "*.xml")
        pdf_paths = _safe_relative_files(xml_dir, "*.pdf")
    else:
        scope = None
        xml_paths = _raw_files(base, ".xml")
        pdf_paths = _raw_files(base, ".pdf")
    if not xml_paths and not structured_input:
        raise ValidationError("Nenhum XML encontrado dentro da pasta autorizada")
    if len(xml_paths) > MAX_XML_FILES:
        raise ValidationError(f"O piloto aceita no máximo {MAX_XML_FILES} XMLs")
    if len(pdf_paths) > MAX_PDF_FILES:
        raise ValidationError(f"O piloto aceita no máximo {MAX_PDF_FILES} PDFs")

    documents: list[dict[str, Any]] = []
    events: list[dict[str, Any]] = []
    file_errors: list[dict[str, Any]] = []
    all_hashes: list[str] = []
    for path in xml_paths:
        parsed_documents, event, error = _parse_xml_file(path)
        if parsed_documents:
            for document in parsed_documents:
                document["folder_direction_hint"] = _direction_hint(path)
                documents.append(document)
            all_hashes.append(parsed_documents[0]["source_hash"])
        elif event:
            events.append(event)
            all_hashes.append(event["source_hash"])
        elif error:
            file_errors.append(error)
            all_hashes.append(error["source_hash"])

    if scope is None:
        scope = _discover_scope(documents, xml_paths)
    _apply_document_rules(documents, events, scope)
    report_identity, identity_findings = _company_report_identity(documents, scope)
    documents_by_key: dict[str, dict[str, Any]] = {}
    for document in documents:
        key = document["document_key"]
        if (
            key
            and document["validity_status"] != "DUPLICATE"
            and key not in documents_by_key
        ):
            documents_by_key[key] = document
    pdf_records = [_parse_pdf_file(path, documents_by_key) for path in pdf_paths]
    all_hashes.extend(record["source_hash"] for record in pdf_records)
    reports, report_hashes = _load_reports(base)
    reconciliations = _reconcile(documents, reports)

    blockers: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = list(identity_findings)
    if not reports:
        warnings.append({"code": "REPORT_MISSING", "ref": None})
    for error in file_errors:
        blockers.append(
            {
                "code": error["status"],
                "ref": error["document_ref"],
                "scope": error.get("analysis_scope", "UNKNOWN"),
            }
        )
    for document in documents:
        if document["scope_status"] == "IN_SCOPE" and document["validity_status"] in {
            "INVALID_STRUCTURE",
            "STATUS_NOT_VERIFIABLE",
        }:
            blockers.append(
                {
                    "code": document["validity_status"],
                    "ref": document["document_ref"],
                    "scope": _analysis_scope(document),
                }
            )
    accepted_reconciliation_statuses = {"MATCHED_VALID", "MATCHED_CANCELLED"}
    for reconciliation in reconciliations:
        if reconciliation["status"] not in accepted_reconciliation_statuses:
            warnings.append(
                {
                    "code": reconciliation["status"],
                    "ref": reconciliation["document_ref"]
                    or reconciliation["report_ref"],
                }
            )
    for document in documents:
        if "FOLDER_DIRECTION_CONFLICT" in document["warnings"]:
            warnings.append(
                {"code": "FOLDER_DIRECTION_CONFLICT", "ref": document["document_ref"]}
            )
    if any(
        "NFSE_STATUS_NOT_EMBEDDED" in document["warnings"] for document in documents
    ):
        warnings.append({"code": "NFSE_STATUS_NOT_EMBEDDED", "ref": None})
    for pdf_record in pdf_records:
        for warning in pdf_record["warnings"]:
            warnings.append({"code": warning, "ref": pdf_record["pdf_ref"]})

    included_documents = [document for document in documents if document["included"]]
    if not included_documents:
        blockers.append(
            {"code": "NO_ELIGIBLE_DOCUMENTS", "ref": None, "scope": "UNKNOWN"}
        )

    blocker_keys = {(item["code"], item["ref"], item["scope"]) for item in blockers}
    blockers = [
        {"code": code, "ref": ref, "scope": scope_code}
        for code, ref, scope_code in sorted(
            blocker_keys, key=lambda item: (item[0], item[1] or "", item[2])
        )
    ]
    warning_keys = {(item["code"], item["ref"]) for item in warnings}
    warnings = [
        {"code": code, "ref": ref}
        for code, ref in sorted(warning_keys, key=lambda item: (item[0], item[1] or ""))
    ]
    scope_authorizations, authorized_scopes, restricted_scopes = (
        _build_scope_authorizations(
            documents,
            file_errors,
            blockers,
            scope["document_families"],
        )
    )
    planning_authorized = bool(authorized_scopes)
    full_documentary_coverage_ready = planning_authorized and not restricted_scopes
    documented_amount = _sum_decimal(
        document["gross_amount"] for document in included_documents
    )
    reported_amount = _sum_decimal(report["gross_amount"] for report in reports)
    matched_report_refs = {
        item["report_ref"]
        for item in reconciliations
        if item["status"] in accepted_reconciliation_statuses
    }
    matched_amount = _sum_decimal(
        report["gross_amount"]
        for report in reports
        if report["report_ref"] in matched_report_refs
    )
    unreconciled_amount = reported_amount - matched_amount

    validation_material = {
        "scope": {
            "entity_ref": scope["entity_ref"],
            "establishment_ref": scope["establishment_ref"],
            "period": scope["period"],
            "document_families": scope["document_families"],
            "analysis_cutoff": scope["analysis_cutoff"],
            "validation_policy": scope["validation_policy"],
            "input_mode": scope["input_mode"],
        },
        "source_hashes": sorted(all_hashes + report_hashes),
    }
    validation_id = (
        "VAL-"
        + hashlib.sha256(
            json.dumps(
                validation_material, sort_keys=True, separators=(",", ":")
            ).encode()
        )
        .hexdigest()[:16]
        .upper()
    )

    excluded_statuses = Counter(
        document["validity_status"]
        for document in documents
        if not document["included"]
    )
    excluded_statuses.update(error["status"] for error in file_errors)
    sanitized_documents = [
        {
            "document_ref": document["document_ref"],
            "source_hash": document["source_hash"],
            "document_type": document["document_type"],
            "validity_status": document["validity_status"],
            "scope_status": document["scope_status"],
            "direction": document["direction"],
            "analysis_scope": _analysis_scope(document),
            "analysis_group": _analysis_group(document),
            "authorized_for_planning": document["included"]
            and _analysis_scope(document) in authorized_scopes,
            "operational_analysis_required": document["included"]
            and _analysis_scope(document) in authorized_scopes,
            "cte_taker_role": document.get("taker_role"),
            "folder_direction_hint": document.get("folder_direction_hint", "UNKNOWN"),
            "emission_period": document["emission_period"],
            "gross_amount": _format_decimal(document["gross_amount"]),
            "item_count": document["item_count"],
            "included": document["included"],
            "reason_codes": sorted(set(document["reason_codes"])),
            "warnings": document["warnings"],
        }
        for document in documents
    ]
    direction_counts = Counter(document["direction"] for document in included_documents)
    document_type_counts = Counter(
        document["document_type"] for document in included_documents
    )
    direction_amounts = {
        direction: _format_decimal(
            _sum_decimal(
                document["gross_amount"]
                for document in included_documents
                if document["direction"] == direction
            )
        )
        for direction in sorted(direction_counts)
    }
    analysis_groups = {}
    for code, definition in ANALYSIS_GROUP_DEFINITIONS.items():
        analysis_scope = _analysis_scope(
            {"document_type": definition["document_types"][0]}
        )
        grouped_documents = [
            document
            for document in included_documents
            if _analysis_group(document) == code
        ]
        detected_group_documents = [
            document
            for document in documents
            if _analysis_group(document) == code
            and document.get("scope_status") == "IN_SCOPE"
        ]
        scope_authorized = analysis_scope in authorized_scopes
        if not detected_group_documents:
            movement_status = "SEM_MOVIMENTACAO"
        elif scope_authorized and grouped_documents:
            movement_status = "COM_MOVIMENTACAO"
        else:
            movement_status = "MOVIMENTACAO_RESTRITA"
        analysis_groups[code] = {
            "label": definition["label"],
            "direction": definition["direction"],
            "analysis_scope": analysis_scope,
            "authorized": scope_authorized,
            "movement_status": movement_status,
            "operational_analysis_required": bool(grouped_documents)
            and scope_authorized,
            "document_types": list(definition["document_types"]),
            "detected_count": len(detected_group_documents),
            "count": len(grouped_documents),
            "gross_amount": _format_decimal(
                _sum_decimal(document["gross_amount"] for document in grouped_documents)
            ),
        }
    authorized_documents = [
        document
        for document in included_documents
        if _analysis_scope(document) in authorized_scopes
    ]
    result = {
        "schema": "br.com.planejamento-reforma-tributaria/document-base-validation",
        "schema_version": "1.7.0",
        "use_case": "UC-001",
        "validation_id": validation_id,
        "status": (
            "DOCUMENT_BASE_BLOCKED"
            if not planning_authorized
            else "DOCUMENT_BASE_READY_WITH_SCOPE_LIMITATIONS"
            if restricted_scopes
            else "DOCUMENT_BASE_READY_WITH_WARNINGS"
            if warnings
            else "DOCUMENT_BASE_READY"
        ),
        "scope": validation_material["scope"],
        "documents": {
            "xml_files_found": len(xml_paths),
            "fiscal_documents_found": len(documents),
            "events_found": len(events),
            "reported": len(reports),
            "included": len(included_documents),
            "excluded": len(documents) - len(included_documents) + len(file_errors),
            "excluded_by_reason": dict(sorted(excluded_statuses.items())),
            "records": sanitized_documents,
            "file_errors": file_errors,
            "document_type_counts": dict(sorted(document_type_counts.items())),
            "direction_counts": dict(sorted(direction_counts.items())),
            "direction_gross_amounts": direction_amounts,
            "analysis_groups": analysis_groups,
        },
        "pdf_evidence": {
            "pdf_files_found": len(pdf_paths),
            "status_counts": dict(
                sorted(Counter(record["status"] for record in pdf_records).items())
            ),
            "keys_detected": sum(record["keys_detected"] for record in pdf_records),
            "identifiers_detected": sum(
                record["identifiers_detected"] for record in pdf_records
            ),
            "matched_document_references": len(
                {
                    document_ref
                    for record in pdf_records
                    for document_ref in record["matched_document_refs"]
                }
            ),
            "records": pdf_records,
        },
        "reconciliation": {
            "status_counts": dict(
                sorted(Counter(item["status"] for item in reconciliations).items())
            ),
            "records": reconciliations,
            "documented_gross_amount": _format_decimal(documented_amount),
            "reported_population_amount": _format_decimal(reported_amount),
            "matched_amount": _format_decimal(matched_amount),
            "unreconciled_amount": _format_decimal(unreconciled_amount),
        },
        "scope_authorizations": scope_authorizations,
        "gates": {
            "reconciliation_ready": bool(reports)
            and all(
                item["status"] in accepted_reconciliation_statuses
                for item in reconciliations
            ),
            "document_analysis_ready": planning_authorized,
            "scope_analysis_ready": planning_authorized,
            "full_documentary_coverage_ready": full_documentary_coverage_ready,
            "authorized_scopes": authorized_scopes,
            "restricted_scopes": restricted_scopes,
            "item_analysis_ready": planning_authorized
            and all(document["item_count"] > 0 for document in authorized_documents),
            "planning_authorized": planning_authorized,
        },
        "blockers": blockers,
        "warnings": warnings,
        "_private_report_context": report_identity,
        "limitations": [
            "A situação atual não foi consultada na autoridade fiscal.",
            "A assinatura digital não foi validada criptograficamente.",
            "Os XMLs não foram validados contra XSD oficial nesta versão do piloto.",
            "NFS-e é validada documentalmente a partir do XML fornecido, sem consulta atual à prefeitura.",
            "CT-e é validado no modelo 57, leiaute 4.00, quando a empresa é emitente ou tomadora identificável.",
            "O UC-001 suporta NF-e, NFC-e, NFS-e ABRASF com InfNfse, CT-e modelo 57 e cancelamentos de NF-e fornecidos na pasta.",
            "CT-e OS, eventos de CT-e e participação sem papel de emitente ou tomador ainda não são suportados.",
            "Validade documental não comprova conformidade tributária ou direito a crédito.",
        ],
        "source_hashes": sorted(all_hashes + report_hashes),
    }
    return result


def _markdown_report(result: dict[str, Any]) -> str:
    documents = result["documents"]
    pdf_evidence = result["pdf_evidence"]
    reconciliation = result["reconciliation"]
    report_identity = result["_private_report_context"]
    lines = [
        "# Relatório de Prontidão da Base Documental",
        "",
        f"- Validação: `{result['validation_id']}`",
        f"- Situação: `{result['status']}`",
        f"- Razão social: {_markdown_escape(report_identity['company_name'])}",
        f"- CNPJ: {_format_cnpj(report_identity['company_taxpayer_id'])}",
        f"- Empresa: `{result['scope']['entity_ref']}`",
        f"- Estabelecimento: `{result['scope']['establishment_ref']}`",
        f"- Competência: `{result['scope']['period']}`",
        f"- Modo de entrada: `{result['scope']['input_mode']}`",
        f"- Data de corte: `{result['scope']['analysis_cutoff']}`",
        "",
        "## Cobertura",
        "",
        f"- XMLs encontrados: {documents['xml_files_found']}",
        f"- Documentos fiscais identificados: {documents['fiscal_documents_found']}",
        f"- Documentos incluídos: {documents['included']}",
        f"- Documentos excluídos: {documents['excluded']}",
        f"- Registros declarados: {documents['reported']}",
        f"- PDFs fiscais encontrados: {pdf_evidence['pdf_files_found']}",
        "",
        "## Autorizações por escopo",
        "",
        "| Escopo | Estado | Autorizado | Incluídos | Excluídos | Bloqueadores |",
        "|---|---|---|---:|---:|---|",
    ]
    for scope_code, authorization in result["scope_authorizations"].items():
        blocker_text = ", ".join(authorization["blocker_codes"]) or "-"
        lines.append(
            f"| `{scope_code}` | `{authorization['status']}` | "
            f"{'sim' if authorization['authorized'] else 'não'} | "
            f"{authorization['included']} | {authorization['excluded']} | "
            f"{blocker_text} |"
        )
    lines.extend(
        [
            "",
            "## Documentos por modelo",
            "",
            "| Modelo | Documentos incluídos |",
            "|---|---:|",
        ]
    )
    for document_type, count in documents["document_type_counts"].items():
        lines.append(f"| `{document_type}` | {count} |")
    lines.extend(
        [
            "",
            "## Entradas e saídas",
            "",
            "| Direção | Documentos | Valor bruto |",
            "|---|---:|---:|",
        ]
    )
    for direction, count in documents["direction_counts"].items():
        lines.append(
            f"| `{direction}` | {count} | {documents['direction_gross_amounts'][direction]} |"
        )
    lines.extend(
        [
            "",
            "## Separação operacional para análise futura",
            "",
            "| Grupo | Escopo | Estado | Análise operacional | Direção | Ocorrências | Incluídos | Valor bruto |",
            "|---|---|---|---|---|---:|---:|---:|",
        ]
    )
    movement_groups = [
        (code, group)
        for code, group in documents["analysis_groups"].items()
        if group["detected_count"] > 0
    ]
    for code, group in movement_groups:
        lines.append(
            f"| `{code}` | `{group['analysis_scope']}` | "
            f"`{group['movement_status']}` | "
            f"{'criar' if group['operational_analysis_required'] else 'não criar'} | "
            f"`{group['direction']}` | {group['detected_count']} | "
            f"{group['count']} | {group['gross_amount']} |"
        )
    if not movement_groups:
        lines.append("| - | - | `SEM_MOVIMENTACAO` | não criar | - | 0 | 0 | 0.00 |")
    without_movement = [
        code
        for code, group in documents["analysis_groups"].items()
        if group["movement_status"] == "SEM_MOVIMENTACAO"
    ]
    lines.extend(
        [
            "",
            "- Sem movimentação: "
            + (", ".join(f"`{code}`" for code in without_movement) or "nenhum grupo"),
        ]
    )
    lines.extend(
        [
            "",
            "## Evidência PDF (DANFE/DACTE/NFS-e)",
            "",
            f"- Chaves detectadas nos PDFs: {pdf_evidence['keys_detected']}",
            f"- Identificadores fiscais detectados: {pdf_evidence['identifiers_detected']}",
            f"- Documentos XML associados: {pdf_evidence['matched_document_references']}",
            "",
            "| Estado | Quantidade |",
            "|---|---:|",
        ]
    )
    for status, count in pdf_evidence["status_counts"].items():
        lines.append(f"| `{status}` | {count} |")
    lines.extend(
        [
            "",
            "## Valores",
            "",
            f"- Valor documentado incluído: {reconciliation['documented_gross_amount']}",
            f"- População declarada: {reconciliation['reported_population_amount']}",
            f"- Valor conciliado: {reconciliation['matched_amount']}",
            f"- Valor não conciliado: {reconciliation['unreconciled_amount']}",
            "",
            "## Conciliação",
            "",
            "| Estado | Quantidade |",
            "|---|---:|",
        ]
    )
    for status, count in reconciliation["status_counts"].items():
        lines.append(f"| `{status}` | {count} |")
    lines.extend(["", "## Bloqueadores", ""])
    if result["blockers"]:
        for blocker in result["blockers"]:
            reference = f" - `{blocker['ref']}`" if blocker["ref"] else ""
            lines.append(
                f"- `{blocker['code']}` - escopo `{blocker['scope']}`{reference}"
            )
    else:
        lines.append("- Nenhum bloqueador identificado.")
    lines.extend(["", "## Avisos de checagem complementar", ""])
    if result["warnings"]:
        for warning in result["warnings"]:
            reference = f" - `{warning['ref']}`" if warning["ref"] else ""
            lines.append(f"- `{warning['code']}`{reference}")
    else:
        lines.append("- Nenhum aviso identificado.")
    lines.extend(["", "## Limitações", ""])
    lines.extend(f"- {item}" for item in result["limitations"])
    lines.extend(
        [
            "",
            "## Decisão",
            "",
            "O planejamento está autorizado em todos os escopos detectados."
            if result["gates"]["full_documentary_coverage_ready"]
            else (
                "A análise pode prosseguir somente nos escopos autorizados; "
                "a cobertura documental integral permanece restrita."
                if result["gates"]["planning_authorized"]
                else "O planejamento permanece bloqueado até a resolução das pendências."
            ),
            "",
        ]
    )
    return "\n".join(lines)


def write_outputs(result: dict[str, Any], output_dir: Path | str) -> tuple[Path, Path]:
    target = Path(output_dir).expanduser().resolve()
    target.mkdir(parents=True, exist_ok=True)
    json_path = target / "validation-result.json"
    markdown_path = target / "relatorio-prontidao-documental.md"
    json_result = {
        key: value for key, value in result.items() if not key.startswith("_private_")
    }
    json_path.write_text(
        json.dumps(json_result, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    markdown_path.write_text(_markdown_report(result), encoding="utf-8")
    return json_path, markdown_path
